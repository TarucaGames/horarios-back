import openpyxl
import locale
from datetime import date, datetime, time, timedelta


class FileAnalyzer:
    semana_diccionario = {
        "lunes": {"index": 0, "inicio": 3, "fin": 7},
        "martes": {"index": 1, "inicio": 8, "fin": 12},
        "miércoles": {"index": 2, "inicio": 13, "fin": 17},
        "jueves": {"index": 3, "inicio": 18, "fin": 22},
        "viernes": {"index": 4, "inicio": 23, "fin": 27},
        "sábado": {"index": 5, "inicio": 28, "fin": 32},
        "domingo": {"index": 6, "inicio": 33, "fin": 37},
    }

    def contar_horas_trabajo(self, path, archivo_excel):
        respuesta = []
        response = {
            "id": None,
            "name": archivo_excel,
            "weeks": [],
            "hasErrors": False,
        }
        # self.obj_to_time(hour=15, minutes=59)
        # Cargar el libro de trabajo
        libro_trabajo = openpyxl.load_workbook(path + archivo_excel)

        # Obtener la lista de nombres de las hojas
        nombres_hojas = libro_trabajo.sheetnames

        # Contar cuántas hojas hay en el libro
        cantidad_hojas = len(nombres_hojas)

        salida_dia_anterior = None
        contador_dias_descanso = 0
        contador_dias_trabajo = 0

        for index in range(cantidad_hojas):
            week_number = int(nombres_hojas[index].split()[1])
            week_dates = self.get_dates_of_week(week_number)
            week = {
                "id": None,
                "name": "",
                "days": [],
                "totalHours": 0,
                "workHours": 0,
                "breakHours": 0,
                "nightHours": 0,
                "errors": [],
                "hasErrors": False,
            }
            week["name"] = nombres_hojas[index].upper()
            # Seleccionar la hoja de trabajo
            hoja_trabajo = libro_trabajo.worksheets[index]

            total_horas_trabajo = 0
            total_horas_nocturnas = 0
            total_horas_descanso = 0

            # Inicializar el contador de horas de trabajo
            # Recorrer el diccionario y llamar a la función para cada día
            for dia, info in self.semana_diccionario.items():
                day = {
                    "id": None,
                    "date": week_dates[info["index"]],
                    "name": dia.upper(),
                    "start": None,
                    "end": None,
                    "isFree": False,
                    "errors": [],
                }
                (
                    horas_trabajo,
                    horas_nocturnas,
                    horas_descanso,
                    entrada,
                    salida,
                    es_feriado,
                ) = self.contar_horas_diarias(hoja_trabajo, info["inicio"], info["fin"])
                if salida is not None and entrada is not None:
                    if contador_dias_descanso == 1:
                        err = "No se respetan las 48hs de días libres"
                        print(f"##! -> {err}")
                        respuesta.append(f"##! -> {err}")
                        day["errors"].append(err)
                    contador_dias_trabajo += 1
                    time_entrada = self.obj_to_time(
                        entrada["columna"] + 4, ((entrada["fila"] - 3) % 5) * 15
                    )
                    time_salida = self.obj_to_time(
                        salida["columna"] + 4, (((salida["fila"] - 3) % 5) + 1) * 15
                    )
                    print(
                        dia.upper()
                        + f" - Entrada: {self.to_string(time_entrada)}"
                        + f" - Salida: {self.to_string(time_salida)}"
                    )
                    respuesta.append(
                        dia.upper()
                        + f" - Entrada: {self.to_string(time_entrada)}"
                        + f" - Salida: {self.to_string(time_salida)}"
                    )
                    day["start"] = self.to_string(time_entrada)
                    day["end"] = self.to_string(time_salida)
                    if salida_dia_anterior is not None:
                        siguiente_entrada = self.get_proxima_entrada(
                            salida_dia_anterior
                        )
                        if siguiente_entrada.time() > time_entrada.time():
                            err = "No se respetan las horas de descanso"
                            print(f"##! -> {err}")
                            respuesta.append(f"##! -> {err}")
                            day["errors"].append(err)
                    salida_dia_anterior = time_salida
                    if contador_dias_trabajo > 7:
                        err = "Más de 7 días de trabajo seguidos"
                        print(f"##! -> {err}")
                        respuesta.append(f"##! -> {err}")
                        day["errors"].append(err)
                    contador_dias_descanso = 0
                    day["type"] = "Holiday" if es_feriado else "Working"
                else:
                    day["isFree"] = True
                    day["type"] = "Free"
                    contador_dias_trabajo = 0
                    contador_dias_descanso += 1
                    salida_dia_anterior = None

                total_horas_trabajo += horas_trabajo
                total_horas_nocturnas += horas_nocturnas
                total_horas_descanso += horas_descanso
                if day["errors"]:
                    week["hasErrors"] = True
                week["days"].append(day)

            # Cerrar el libro de trabajo
            libro_trabajo.close()
            print(nombres_hojas[index].upper())
            respuesta.append(nombres_hojas[index].upper())
            print(
                f"Total horas: {total_horas_trabajo + total_horas_descanso} \nHoras recepción: {total_horas_trabajo} ({total_horas_nocturnas} son nocturas) \nHoras descanso: {total_horas_descanso}"
            )
            respuesta.append(
                f"Total horas: {total_horas_trabajo + total_horas_descanso} \nHoras recepción: {total_horas_trabajo} ({total_horas_nocturnas} son nocturas) \nHoras descanso: {total_horas_descanso}"
            )
            print("=======")
            respuesta.append("=======")
            week["breakHours"] = total_horas_descanso
            week["workHours"] = total_horas_trabajo
            week["nightHours"] = total_horas_nocturnas
            week["totalHours"] = total_horas_trabajo + total_horas_descanso
            if week["hasErrors"]:
                response["hasErrors"] = True
            response["weeks"].append(week)
        return respuesta, response

    def contar_horas_diarias(self, hoja_trabajo, inicio, fin):
        # Inicializar el contador de horas de trabajo
        horas_trabajo = 0
        horas_nocturnas = 0
        horas_descanso = 0
        salida = None
        entrada = None

        celda_referencia = hoja_trabajo.cell(row=inicio, column=3)
        es_feriado = celda_referencia.fill.start_color.index == "FFD0CECE"
        # Iterar sobre las filas desde la fila 4 hasta la 6 y las columnas desde la C hasta la S
        for columna in range(3, 20):
            for fila in range(inicio, fin):  # Columnas de la C a la S
                celda = hoja_trabajo.cell(row=fila, column=columna)
                if celda.value == "X":
                    horas_descanso += 0.25
                if celda.value == "R":
                    # Si la celda tiene una 'R', se suma 0.25 horas al contador
                    salida = {"fila": fila, "columna": columna}
                    horas_trabajo += 0.25
                    if columna >= 18:
                        horas_nocturnas += 0.25
                    if entrada is None:
                        entrada = {"fila": fila, "columna": columna}
        return (
            horas_trabajo,
            horas_nocturnas,
            horas_descanso,
            entrada,
            salida,
            es_feriado,
        )

    def obj_to_time(self, hour, minutes):
        return datetime.combine(date.today(), time(hour=hour, minute=0)) + timedelta(
            minutes=minutes
        )

    def to_string(self, time):
        return time.strftime("%H:%M")

    def get_proxima_entrada(self, salida_dia_anterior):
        # devolver cuando debería ser la proxima entrada
        return salida_dia_anterior + timedelta(hours=12)

    def analyze(self):
        return "a"

    def get_dates_of_week(self, week_number):
        # Set the locale to Spanish
        locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")

        # Determine the first day of the week
        first_day = datetime.strptime(f"2024-W{week_number}-1", "%Y-W%W-%w").date()

        # Calculate the dates for the entire week
        dates_of_week = [first_day + timedelta(days=i) for i in range(7)]

        # Format dates in Spanish
        formatted_dates = [date.strftime("%d-%m") for date in dates_of_week]

        return formatted_dates
