import openpyxl
import locale
from datetime import date, datetime, time, timedelta

from models import CreateShiftModel, ShiftModel, EmployeeCreate, ShiftCreate
from services.shift_service import ShiftService


class FileReader:
    semana_diccionario = {
        "lunes": {"index": 0, "inicio": 3, "fin": 7},
        "martes": {"index": 1, "inicio": 8, "fin": 12},
        "miércoles": {"index": 2, "inicio": 13, "fin": 17},
        "jueves": {"index": 3, "inicio": 18, "fin": 22},
        "viernes": {"index": 4, "inicio": 23, "fin": 27},
        "sábado": {"index": 5, "inicio": 28, "fin": 32},
        "domingo": {"index": 6, "inicio": 33, "fin": 37},
    }

    weekdays_dict = {
        "lunes": {"index": 0, "column_start": 4, "column_end": 8},
        "martes": {"index": 1, "column_start": 8, "column_end": 12},
        "miércoles": {"index": 2, "column_start": 12, "column_end": 16},
        "jueves": {"index": 3, "column_start": 16, "column_end": 20},
        "viernes": {"index": 4, "column_start": 20, "column_end": 24},
        "sábado": {"index": 5, "column_start": 24, "column_end": 27},
        "domingo": {"index": 6, "column_start": 27, "column_end": 30},
    }

    def read(self, path, excel_file, employee_service, shift_service):

        employees = employee_service.get_all()
        employees_dict = {employee.code: employee.id for employee in employees}
        print(employees_dict.get("C"))
        # load file
        workbook = openpyxl.load_workbook(path + excel_file)

        # get sheet names
        sheet_names = workbook.sheetnames

        # sheet quantity
        sheet_quantity = len(sheet_names)

        for index in range(1):
            sheet = workbook.worksheets[index]
            week_number = self.get_week_number(sheet)
            employees_dict = self.read_employees(
                sheet, employees_dict, employee_service
            )
            print(employees_dict)
            sheet_info = self.read_sheet(
                sheet, week_number, employees_dict, shift_service
            )
        return True

    def read_employees(self, sheet, employee_list: dict, service):
        row_code = 26
        column_code = 32
        column_name = 34
        # Iterate over the column until a cell is None
        for cell in sheet.iter_rows(
            min_row=row_code, min_col=column_code, max_col=column_code
        ):
            code = cell[0].value
            if code is None:
                break
            if code not in employee_list:
                # create
                name = sheet.cell(cell[0].row, column_name).value
                employee_data = EmployeeCreate(**{"name": name, "code": code})
                employee = service.create(employee_data)
                print(f"created {employee.name}")
                employee_list[code] = employee.id
        return employee_list

    def read_sheet(
        self, sheet, week_number, employees_dict, shift_service: ShiftService
    ):
        # delete previously created shifts. maybe it should be done at the end
        shift_service.delete(week_number)

        # read days data
        week_dates = self.get_dates_of_week(week_number)
        for day, info in self.weekdays_dict.items():
            print(day)
            self.get_daily_shifts(
                sheet,
                info["column_start"],
                info["column_end"],
                week_number,
                week_dates[info["index"]],
                employees_dict,
                shift_service,
            )
        return True

    def get_week_number(self, sheet):
        return int(sheet.cell(14, 34).value)

    def get_daily_shifts(
        self,
        sheet,
        column_start,
        column_end,
        week_number,
        date,
        employees_dict,
        shift_service: ShiftService,
    ):
        # Inicializar el contador de horas de trabajo
        horas_trabajo = 0
        horas_nocturnas = 0
        horas_descanso = 0
        salida = None
        entrada = None

        celda_referencia = sheet.cell(row=7, column=column_start)
        for column in range(column_start, column_end):
            previous = None
            current = None
            for row in range(8, 74):
                cell_value = sheet.cell(row, column).value
                if cell_value != previous:
                    if previous:
                        end_time = self.obj_to_time(hour=7, minutes=(row - 8) * 15)
                        current.end_time = self.to_string(end_time)
                        shift_service.create(current)
                        # print(f"Termina {previous}: [{column}][{row-1}]")
                        print(current)
                    if cell_value:
                        start_time = self.obj_to_time(hour=7, minutes=(row - 8) * 15)
                        data = {
                            "employee_id": employees_dict.get(cell_value),
                            "start_time": self.to_string(start_time),
                            "end_time": None,
                            "week": week_number,
                            "type": 1,
                            "date": str(date),
                        }
                        current = ShiftCreate(**data)
                        # print(f"Empieza {cell_value}: [{column}][{row}]")
                previous = cell_value

        """ es_feriado = celda_referencia.fill.start_color.index == "FFD0CECE"
        # Iterar sobre las filas desde la fila 4 hasta la 6 y las columnas desde la C hasta la S
        for columna in range(3, 20):
            for fila in range(column_start, column_end):  # Columnas de la C a la S
                celda = sheet.cell(row=fila, column=columna)
                if celda.value == "X":
                    horas_descanso += 0.25
                if celda.value == "R":
                    # Si la celda tiene una 'R', se suma 0.25 horas al contador
                    salida = {"fila": fila, "columna": columna}
                    horas_trabajo += 0.25
                    if columna >= 18:
                        horas_nocturnas += 0.25
                    if entrada is None:
                        entrada = {"fila": fila, "columna": columna}
        return (
            horas_trabajo,
            horas_nocturnas,
            horas_descanso,
            entrada,
            salida,
            es_feriado,
        ) """

    def read_single(
        self, path, excel_file, employee_service, shift_service, employee_id
    ):
        respuesta = []
        response = {
            "id": None,
            "name": excel_file,
            "weeks": [],
            "hasErrors": False,
        }
        employee = employee_service.get(employee_id)
        print(employee.name)
        # Open workbook
        workbook = openpyxl.load_workbook(path + excel_file)
        # Get sheet names
        sheet_names = workbook.sheetnames
        # Count sheets
        sheet_quantity = len(sheet_names)
        salida_dia_anterior = None
        contador_dias_descanso = 0
        contador_dias_trabajo = 0

        for index in range(sheet_quantity):
            week_number = int(sheet_names[index].split()[1])
            week_dates = self.get_dates_of_week(week_number)
            week = {
                "id": None,
                "name": "",
                "days": [],
                "totalHours": 0,
                "workHours": 0,
                "breakHours": 0,
                "nightHours": 0,
                "errors": [],
                "hasErrors": False,
            }
            week["name"] = sheet_names[index].upper()
            # Seleccionar la hoja de trabajo
            hoja_trabajo = workbook.worksheets[index]

            total_horas_trabajo = 0
            total_horas_nocturnas = 0
            total_horas_descanso = 0

            # Inicializar el contador de horas de trabajo
            # Recorrer el diccionario y llamar a la función para cada día
            for dia, info in self.semana_diccionario.items():
                date = week_dates[info["index"]]
                day = {
                    "id": None,
                    "date": date,
                    "name": dia.upper(),
                    "start": None,
                    "end": None,
                    "isFree": False,
                    "errors": [],
                }
                (
                    horas_trabajo,
                    horas_nocturnas,
                    horas_descanso,
                    entrada,
                    salida,
                    es_feriado,
                ) = self.contar_horas_diarias(hoja_trabajo, info["inicio"], info["fin"])
                if salida is not None and entrada is not None:
                    if contador_dias_descanso == 1:
                        err = "No se respetan las 48hs de días libres"
                        print(f"##! -> {err}")
                        respuesta.append(f"##! -> {err}")
                        day["errors"].append(err)
                    contador_dias_trabajo += 1
                    time_entrada = self.obj_to_time(
                        entrada["columna"] + 4, ((entrada["fila"] - 3) % 5) * 15
                    )
                    time_salida = self.obj_to_time(
                        salida["columna"] + 4, (((salida["fila"] - 3) % 5) + 1) * 15
                    )
                    print(
                        dia.upper()
                        + f" - Entrada: {self.to_string(time_entrada)}"
                        + f" - Salida: {self.to_string(time_salida)}"
                    )
                    respuesta.append(
                        dia.upper()
                        + f" - Entrada: {self.to_string(time_entrada)}"
                        + f" - Salida: {self.to_string(time_salida)}"
                    )
                    day["start"] = self.to_string(time_entrada)
                    day["end"] = self.to_string(time_salida)
                    if salida_dia_anterior is not None:
                        siguiente_entrada = self.get_proxima_entrada(
                            salida_dia_anterior
                        )
                        if siguiente_entrada.time() > time_entrada.time():
                            err = "No se respetan las horas de descanso"
                            print(f"##! -> {err}")
                            respuesta.append(f"##! -> {err}")
                            day["errors"].append(err)
                    salida_dia_anterior = time_salida
                    if contador_dias_trabajo > 7:
                        err = "Más de 7 días de trabajo seguidos"
                        print(f"##! -> {err}")
                        respuesta.append(f"##! -> {err}")
                        day["errors"].append(err)
                    contador_dias_descanso = 0
                    day["type"] = "Holiday" if es_feriado else "Working"
                    data = {
                        "employee_id": employee_id,
                        "start_time": self.to_string(time_entrada),
                        "end_time": self.to_string(time_salida),
                        "week": week_number,
                        "type": 1,
                        "date": str(date),
                    }
                    current = ShiftCreate(**data)
                    shift_service.create(current)
                    print(current)
                else:
                    day["isFree"] = True
                    day["type"] = "Free"
                    contador_dias_trabajo = 0
                    contador_dias_descanso += 1
                    salida_dia_anterior = None

                total_horas_trabajo += horas_trabajo
                total_horas_nocturnas += horas_nocturnas
                total_horas_descanso += horas_descanso
                if day["errors"]:
                    week["hasErrors"] = True
                week["days"].append(day)

            # Cerrar el libro de trabajo
            workbook.close()
            print(sheet_names[index].upper())
            respuesta.append(sheet_names[index].upper())
            print(
                f"Total horas: {total_horas_trabajo + total_horas_descanso} \nHoras recepción: {total_horas_trabajo} ({total_horas_nocturnas} son nocturas) \nHoras descanso: {total_horas_descanso}"
            )
            respuesta.append(
                f"Total horas: {total_horas_trabajo + total_horas_descanso} \nHoras recepción: {total_horas_trabajo} ({total_horas_nocturnas} son nocturas) \nHoras descanso: {total_horas_descanso}"
            )
            print("=======")
            respuesta.append("=======")
            week["breakHours"] = total_horas_descanso
            week["workHours"] = total_horas_trabajo
            week["nightHours"] = total_horas_nocturnas
            week["totalHours"] = total_horas_trabajo + total_horas_descanso
            if week["hasErrors"]:
                response["hasErrors"] = True
            response["weeks"].append(week)
        return response

    def contar_horas_trabajo(self, path, archivo_excel):
        respuesta = []
        response = {
            "id": None,
            "name": archivo_excel,
            "weeks": [],
            "hasErrors": False,
        }
        # self.obj_to_time(hour=15, minutes=59)
        # Cargar el libro de trabajo
        libro_trabajo = openpyxl.load_workbook(path + archivo_excel)

        # Obtener la lista de nombres de las hojas
        nombres_hojas = libro_trabajo.sheetnames

        # Contar cuántas hojas hay en el libro
        cantidad_hojas = len(nombres_hojas)

        salida_dia_anterior = None
        contador_dias_descanso = 0
        contador_dias_trabajo = 0

        for index in range(cantidad_hojas):
            week_number = int(nombres_hojas[index].split()[1])
            week_dates = self.get_dates_of_week(week_number)
            week = {
                "id": None,
                "name": "",
                "days": [],
                "totalHours": 0,
                "workHours": 0,
                "breakHours": 0,
                "nightHours": 0,
                "errors": [],
                "hasErrors": False,
            }
            week["name"] = nombres_hojas[index].upper()
            # Seleccionar la hoja de trabajo
            hoja_trabajo = libro_trabajo.worksheets[index]

            total_horas_trabajo = 0
            total_horas_nocturnas = 0
            total_horas_descanso = 0

            # Inicializar el contador de horas de trabajo
            # Recorrer el diccionario y llamar a la función para cada día
            for dia, info in self.semana_diccionario.items():
                day = {
                    "id": None,
                    "date": week_dates[info["index"]],
                    "name": dia.upper(),
                    "start": None,
                    "end": None,
                    "isFree": False,
                    "errors": [],
                }
                (
                    horas_trabajo,
                    horas_nocturnas,
                    horas_descanso,
                    entrada,
                    salida,
                    es_feriado,
                ) = self.contar_horas_diarias(hoja_trabajo, info["inicio"], info["fin"])
                if salida is not None and entrada is not None:
                    if contador_dias_descanso == 1:
                        err = "No se respetan las 48hs de días libres"
                        print(f"##! -> {err}")
                        respuesta.append(f"##! -> {err}")
                        day["errors"].append(err)
                    contador_dias_trabajo += 1
                    time_entrada = self.obj_to_time(
                        entrada["columna"] + 4, ((entrada["fila"] - 3) % 5) * 15
                    )
                    time_salida = self.obj_to_time(
                        salida["columna"] + 4, (((salida["fila"] - 3) % 5) + 1) * 15
                    )
                    print(
                        dia.upper()
                        + f" - Entrada: {self.to_string(time_entrada)}"
                        + f" - Salida: {self.to_string(time_salida)}"
                    )
                    respuesta.append(
                        dia.upper()
                        + f" - Entrada: {self.to_string(time_entrada)}"
                        + f" - Salida: {self.to_string(time_salida)}"
                    )
                    day["start"] = self.to_string(time_entrada)
                    day["end"] = self.to_string(time_salida)
                    if salida_dia_anterior is not None:
                        siguiente_entrada = self.get_proxima_entrada(
                            salida_dia_anterior
                        )
                        if siguiente_entrada.time() > time_entrada.time():
                            err = "No se respetan las horas de descanso"
                            print(f"##! -> {err}")
                            respuesta.append(f"##! -> {err}")
                            day["errors"].append(err)
                    salida_dia_anterior = time_salida
                    if contador_dias_trabajo > 7:
                        err = "Más de 7 días de trabajo seguidos"
                        print(f"##! -> {err}")
                        respuesta.append(f"##! -> {err}")
                        day["errors"].append(err)
                    contador_dias_descanso = 0
                    day["type"] = "Holiday" if es_feriado else "Working"
                else:
                    day["isFree"] = True
                    day["type"] = "Free"
                    contador_dias_trabajo = 0
                    contador_dias_descanso += 1
                    salida_dia_anterior = None

                total_horas_trabajo += horas_trabajo
                total_horas_nocturnas += horas_nocturnas
                total_horas_descanso += horas_descanso
                if day["errors"]:
                    week["hasErrors"] = True
                week["days"].append(day)

            # Cerrar el libro de trabajo
            libro_trabajo.close()
            print(nombres_hojas[index].upper())
            respuesta.append(nombres_hojas[index].upper())
            print(
                f"Total horas: {total_horas_trabajo + total_horas_descanso} \nHoras recepción: {total_horas_trabajo} ({total_horas_nocturnas} son nocturas) \nHoras descanso: {total_horas_descanso}"
            )
            respuesta.append(
                f"Total horas: {total_horas_trabajo + total_horas_descanso} \nHoras recepción: {total_horas_trabajo} ({total_horas_nocturnas} son nocturas) \nHoras descanso: {total_horas_descanso}"
            )
            print("=======")
            respuesta.append("=======")
            week["breakHours"] = total_horas_descanso
            week["workHours"] = total_horas_trabajo
            week["nightHours"] = total_horas_nocturnas
            week["totalHours"] = total_horas_trabajo + total_horas_descanso
            if week["hasErrors"]:
                response["hasErrors"] = True
            response["weeks"].append(week)
        return respuesta, response

    def contar_horas_diarias(self, hoja_trabajo, inicio, fin):
        # Inicializar el contador de horas de trabajo
        horas_trabajo = 0
        horas_nocturnas = 0
        horas_descanso = 0
        salida = None
        entrada = None

        celda_referencia = hoja_trabajo.cell(row=inicio, column=3)
        es_feriado = celda_referencia.fill.start_color.index == "FFD0CECE"
        # Iterar sobre las filas desde la fila 4 hasta la 6 y las columnas desde la C hasta la S
        for columna in range(3, 20):
            for fila in range(inicio, fin):  # Columnas de la C a la S
                celda = hoja_trabajo.cell(row=fila, column=columna)
                if celda.value == "X":
                    horas_descanso += 0.25
                if celda.value == "R":
                    # Si la celda tiene una 'R', se suma 0.25 horas al contador
                    salida = {"fila": fila, "columna": columna}
                    horas_trabajo += 0.25
                    if columna >= 18:
                        horas_nocturnas += 0.25
                    if entrada is None:
                        entrada = {"fila": fila, "columna": columna}
        return (
            horas_trabajo,
            horas_nocturnas,
            horas_descanso,
            entrada,
            salida,
            es_feriado,
        )

    def obj_to_time(self, hour, minutes):
        return datetime.combine(date.today(), time(hour=hour, minute=0)) + timedelta(
            minutes=minutes
        )

    def to_string(self, time):
        return time.strftime("%H:%M")

    def get_proxima_entrada(self, salida_dia_anterior):
        # devolver cuando debería ser la proxima entrada
        return salida_dia_anterior + timedelta(hours=12)

    def analyze(self):
        return "a"

    def get_dates_of_week(self, week_number):
        # Set the locale to Spanish
        locale.setlocale(locale.LC_TIME, "es_ES.UTF-8")

        # Determine the first day of the week
        first_day = datetime.strptime(f"2024-W{week_number}-1", "%Y-W%W-%w").date()

        # Calculate the dates for the entire week
        dates_of_week = [first_day + timedelta(days=i) for i in range(7)]

        # Format dates in Spanish
        # formatted_dates = [date.strftime("%d-%m") for date in dates_of_week]
        formatted_dates = [date.strftime("%Y-%m-%d") for date in dates_of_week]

        return formatted_dates
